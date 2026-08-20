import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# Cargar el archivo Excel
archivo = 'aspersion.xlsx'
wb = openpyxl.load_workbook(archivo, data_only=False)
ws = wb.active

# Información de salida
salida = []
salida.append(f"Análisis del archivo: {archivo}")
salida.append(f"Hoja activa: {ws.title}")
salida.append(f"Fecha de análisis: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
salida.append("=" * 100)
salida.append("")

# Iterar sobre todas las filas
celdas_count = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            celdas_count += 1
            # Obtener información de la celda
            referencia = cell.coordinate
            valor = cell.value
            tipo_dato = type(valor).__name__
            
            # Formatear el tipo de dato para legibilidad
            if tipo_dato == "str":
                tipo_formateado = "Texto"
            elif tipo_dato == "int":
                tipo_formateado = "Número Entero"
            elif tipo_dato == "float":
                tipo_formateado = "Número Decimal"
            elif tipo_dato == "bool":
                tipo_formateado = "Booleano"
            elif tipo_dato == "datetime":
                tipo_formateado = "Fecha/Hora"
            else:
                tipo_formateado = tipo_dato
            
            # Crear línea de salida
            linea = f"Celda: {referencia:6} | Tipo: {tipo_formateado:20} | Valor: {str(valor)[:80]}"
            salida.append(linea)

salida.append("")
salida.append("=" * 100)
salida.append(f"Total de celdas con valores: {celdas_count}")

# Imprimir en pantalla
for linea in salida:
    print(linea)

# Guardar en archivo de texto
nombre_archivo_salida = "mapeo_aspersion.txt"
with open(nombre_archivo_salida, 'w', encoding='utf-8') as f:
    for linea in salida:
        f.write(linea + "\n")

print(f"\nArchivo guardado en: {nombre_archivo_salida}")
